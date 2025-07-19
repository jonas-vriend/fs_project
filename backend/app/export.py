import csv
import openpyxl as xl
import openpyxl.styles as style
import openpyxl.utils as utils
from .models import  FinancialStatement

# Define shared styles
BLACK_FILL = style.PatternFill(fill_type="solid", start_color="FF000000", end_color="FF000000")
YELLOW_FILL = style.PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
HEADER_FONT = style.Font(name="Helvetica Neue", size=10, color="FFFFFF", bold=True)
BLUE_FONT = style.Font(name="Helvetica Neue", size=8, color="0070c0")
BLACK_FONT = style.Font(name="Helvetica Neue", size=8, color="000000")
RED_ITALIC_FONT = style.Font(name="Helvetica Neue", size=8, color="FF0000", italic=True)
LABEL_FONT = style.Font(name="Helvetica Neue", size=8)

THIN_LINE = style.Side(border_style="thin", color="000000")
DOUBLE_LINE = style.Side(border_style="double", color="000000")
SUBTOTAL_BORDER = style.Border(top=THIN_LINE)
TOTAL_BORDER = style.Border(top=THIN_LINE, bottom=DOUBLE_LINE)


def export_fs_as_csv(fs: FinancialStatement, export_filename):
        """
        Exports FinancialStatement object to CSV
        """
        export_filename = export_filename + ".csv"
        all_years = []
        seen = set()
        for item in fs.lines:
            for year in item.data.keys():
                if year not in seen:
                    seen.add(year)
                    all_years.append(year)

        with open(export_filename, "w", newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Label"] + all_years)

            for item in fs.lines:
                row = [item.label]
                for year in all_years:
                    value = item.data.get(year, "")
                    row.append(value)
                writer.writerow(row)

def export_fs_as_xlsx(fs: FinancialStatement, export_filename):
    export_filename = export_filename + ".xlsx"
    wb = xl.Workbook()
    del wb['Sheet']
    fs_type = fs.get_type()
    ws_new = wb.create_sheet(title=fs_type)
    ws_new.sheet_view.showGridLines = False
    ws_new.column_dimensions["A"].width = 1


    # Run formatting
    lines = fs.get_lines()
    years = fs.get_years()

    format_line_items(ws_new, years, lines, fs_type)
    build_header(ws_new, years, fs)

    # Save to file
    output_file = "formatted_output.xlsx"
    wb.save(output_file)
    print(f"Excel file saved as: {output_file}")


def format_line_items(ws_new, years, line_items, fs_type, start_row=4, start_col=2): #TODO: currently only supports BS

    LABEL_COL_SCALE_FACTOR = 0.8
    s_type_2 = []

    max_label_length = 0
    for i, line in enumerate(line_items):
        row = start_row + i
        label, values, dollar_sign, indent_level, summing_type, summing_range = line.get_all()
        if len(label) > max_label_length:
            max_label_length = len(label)

        # Add label cell
        cell = ws_new.cell(row=row, column=start_col, value=label)
        cell.font = LABEL_FONT

        # Set alignment: center if all caps heading, else indent
        cleaned = label.replace(":", "").replace(" ", "").replace("'", "")
        if cleaned.isupper() and cleaned.isalpha() and not line.get_data():
            cell.alignment = style.Alignment(horizontal="center", indent=0)
        else:
            cell.alignment = style.Alignment(horizontal="left", indent=indent_level)

        if summing_type == 3:
            s_type_2.append(line_items.index(line))

        # Add values
        for j, year in enumerate(years):
            val = values.get(year, '')
            col = start_col + 1 + j
            # Hard code blue font
            if summing_type == 0:
                val_cell = ws_new.cell(row=row, column=col, value=val)
                val_cell.font = BLUE_FONT

            # totals and subtotals given summing formulas, black font, borders
            else:
                if summing_range:
                    included_rows = [idx + start_row for idx in summing_range]
                    col_letter = utils.get_column_letter(col)
                    formula = f"=SUM({','.join(f'{col_letter}{r}' for r in included_rows)})"
                    val = formula
                val_cell = ws_new.cell(row=row, column=col, value=val)
                val_cell.font = BLACK_FONT

                if summing_type in (1,2):
                    val_cell.border = SUBTOTAL_BORDER
                elif summing_type == 3:
                    val_cell.border = TOTAL_BORDER
            # Dollar signs assigned to end and beinning of both A and L + SE
            if dollar_sign:
                val_cell.number_format = '_("$"* #,##0_);_("$"* (#,##0)'
            else:
                val_cell.number_format = '#,##0;(#,##0)'

    # Dynamicallty set label width based on max label length
    col_letter = utils.get_column_letter(start_col)
    ws_new.column_dimensions[col_letter].width = max_label_length * LABEL_COL_SCALE_FACTOR

    # Add balance check
    if fs_type == "BALANCE_SHEET" and len(s_type_2) == 2:
        ta = s_type_2[0] + start_row
        tlse = s_type_2[1] + start_row
        row += 2
        balance_check = ws_new.cell(row=row, column=start_col, value='Balance Check')
        balance_check.font = RED_ITALIC_FONT
        for i, year in enumerate(years):
            col = start_col + 1 + i
            col_letter = utils.get_column_letter(col)
            formula = f"={col_letter}{tlse} - {col_letter}{ta}"
            balance_val = ws_new.cell(row=row, column=col)
            balance_val.value = formula
            balance_val.font = RED_ITALIC_FONT


def build_header(ws_new, years, fs: FinancialStatement, start_row=1, start_col=1):
        # Fill black background
        HEADER_LENGTH = 3
        for i in range(HEADER_LENGTH):
            for j in range(len(years) + 2):
                row = start_row + i
                col = start_col + j
                cell = ws_new.cell(row=row, column=col)
                cell.fill = BLACK_FILL

        # Add Financial Statement type
        row = start_row + 1
        fs_type = fs.get_type()
        label = 'Consolidated Balance Sheets' if fs_type == 'BALANCE_SHEET' else 'Consolidated Income Statements'
        type_cell = ws_new.cell(row=row, column=start_col + 1, value=label)
        type_cell.font = HEADER_FONT

        # Add Years
        for i, year in enumerate(years):
            row = start_row + 2
            col = start_col + 2 + i
            year_cell = ws_new.cell(row=row, column=col, value=year)
            year_cell.alignment = style.Alignment(horizontal="center")
            year_cell.font = HEADER_FONT