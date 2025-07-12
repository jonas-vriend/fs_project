import openpyxl as xl
import openpyxl.styles as style
import openpyxl.utils as utils
import fs_app as fs

# copying main in this to access fs for debugging purposes. Will delete later
def main(debug=False, use_cache=True):
    """
    Orchestrates the pipeline:
    - Load/caches OCR
    - Extracts structure and values
    - Builds financial statement object
    """
    ocr_output, processed_img, underscore_coords = fs.get_data(use_cache, debug)
    col_coords, lines = fs.preprocess_text(ocr_output, debug)
    if underscore_coords:
        fs.add_underscore_zeros(lines, col_coords, underscore_coords, debug)
    fs.debug_output(ocr_output, processed_img, col_coords, val_x_thresh=75, debug=False)
    fs.add_indentation(lines)
    merged_lines = fs.merge_lines(lines, debug)
    completed = fs.build_fs(col_coords, merged_lines, debug)
    return completed

new_fs = main()
lines = new_fs.get_lines()
wb = xl.Workbook()

del wb['Sheet']

sheet_name = new_fs.get_type()
ws_new = wb.create_sheet(title=sheet_name)


def format_line_items(line_items, start_row=4, start_col=2):
    years = new_fs.get_years()
    for i, line in enumerate(line_items):
        row = start_row + i
        label, values, dollar_sign, indent_level = line.get_all()

        # Add labels 
        cell = ws_new.cell(row=row, column=start_col, value=label)

        # Set indentation. If header is uppercase, center.
        if label.isupper() and label.replace(" ", "").isalpha():
            alignment = style.Alignment(horizontal="center", indent=0)
        else:
            alignment = style.Alignment(horizontal="left", indent=indent_level)

        cell.alignment = alignment

        # Add values corresponding to each year
        for j, year in enumerate(years):
            val = values.get(year, '')
            val_cell = ws_new.cell(row=row, column=start_col + 1 + j, value=val)
            val_cell.font = style.Font(color="0000FF")  # Blue font default for hardcoded vals
            
            # Format dollar signs and negagive numbers
            if isinstance(val, int):
                if dollar_sign:
                    val_cell.number_format = '$    #,##0;$    (#,##0)'
                else:
                    val_cell.number_format = '#,##0;(#,##0)'

            else:
                val_cell.number_format = 'General'
        
    # Add balance check
    row += 2
    balance_check = ws_new.cell(row=row, column=start_col, value='Balance Check')
    balance_check.font = style.Font(color="FF0000", italic=True)

    for i, year in enumerate(years):
        column = start_col + 1 + i
        col_letter = utils.get_column_letter(column)
        check_formula = f"={col_letter}35 - {col_letter}15"  # Hardcoded but need to change once i make detection of summing lines dynamic
        balance_val = ws_new.cell(row=row, column=column)
        balance_val.value = check_formula
        balance_val.font = style.Font(color="FF0000", italic=True)


            
format_line_items(lines)
output_file = "formatted_output.xlsx"
wb.save(output_file)
print(f"Excel file saved as: {output_file}")    

        